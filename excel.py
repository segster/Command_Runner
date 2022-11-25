#open module for Excel
import openpyxl

#open Excel book
inv_file = openpyxl.load_workbook("inventory.xlsx")

#Assign variable to Exel sheet
product_list = inv_file["Sheet1"]

#create 3 empty ditionary
products_per_supplier = {}
total_value_per_supplier = {}
products_under_10 = {}

#test for max row function
#print (product_list.max_row)


#in each row of worksheet - assigh name of supplier
#in column 4 to variable supp_name. start range at 2 because sheet has headers.

for product_row in range(2, product_list.max_row + 1):
    #supp name = name in cell D2
    supp_name = product_list.cell(product_row, 4).value
    #print (product_list.cell(product_row, 4))
    #set VARS based on CELL values
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 6)

    #print (supp_name)
    #calculation number of products per  supplier
    #if supplier already exists in Dictionary, add product
    if supp_name in products_per_supplier:
        #using GET function from key initalise with current product number
        #from the Dictionary
        current_num_products = products_per_supplier.get(supp_name)
        #print ("**", products_per_supplier.get(supp_name))
        #print (type (products_per_supplier.get(supp_name)))
        #print  (products_per_supplier)
        #print ("#", current_num_products)
        #increment produst per supplier count by 1
        products_per_supplier[supp_name] = current_num_products + 1
    #if supplier does not exist in Dictionary - add it
    else:
        print ("adding a new supplier", supp_name)
        products_per_supplier[supp_name] = 1
        #products_per_supplier["simon"] = 1
        #print (products_per_supplier)

    #calculation total value of inventory per suppplier
    #check if supplier currently exist in Dictionary
    if supp_name in total_value_per_supplier:
        #VAR current_total_value is equal to value based on supplier key
        current_total_value = total_value_per_supplier.get(supp_name)
        total_value_per_supplier[supp_name] = current_total_value + inventory * price
        #print (total_value_per_supplier[supp_name])
    else:
        total_value_per_supplier[supp_name] = inventory * price

    # products under 10
    if inventory <50:
        products_under_10[product_num] = inventory

    #add value for total inventory price
    inventory_price.value = inventory * price + 10

print (products_under_10)


inv_file.save("inventory_2.xlsx")



#print (total_value_per_supplier)
#print (products_per_supplier)

#print (products_per_supplier[supp_name])

'''
for supp, prod in sorted(products_per_supplier.items()):
    #for s, total in total_value_per_supplier.items():
        print (f"The supplier {supp.title()} has {prod} items in the inventory. ")'''

#ADD column to file




